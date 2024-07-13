from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from wind_tunnel_testing import features
from wind_tunnel_testing.config import RAW_DATA_DIR, REPORTS_DIR
from wind_tunnel_testing.dataset import load_ml_test_data

XLSX_DATA_FILENAME = "ML_test"


def load_workbook_safe(file_path):
    """
    Load a workbook safely with error handling.

    Args:
        file_path (str): Path to the Excel file.

    Returns:
        Workbook: The loaded workbook.

    Raises:
        FileNotFoundError: If the file does not exist.
        InvalidFileException: If the file is not a valid Excel file.
    """
    try:
        return load_workbook(file_path)
    except InvalidFileException:
        print(f"Error: The file {file_path} is invalid or corrupted.")
        raise
    except FileNotFoundError:
        print(f"Error: The file {file_path} was not found.")
        raise


def process_workbook(sheet_name, data_to_write, column_map, start_row=2):
    """
    Process the workbook by writing data to a specific sheet and columns.

    Args:
        sheet_name (str): The name of the sheet to write data.
        data_to_write (list of lists): The data to be written.
        column_map (list): List of column letters corresponding to the data.
        start_row (int): The starting row for writing data.
    """
    report = REPORTS_DIR / f"{XLSX_DATA_FILENAME}.xlsx"
    wb = load_workbook_safe(report)
    ws = wb[sheet_name]
    for index, data in enumerate(data_to_write):
        for col_letter, value in zip(column_map, data):
            cell = f"{col_letter}{start_row + index}"
            ws[cell] = value
    wb.save(report)


def surrogate_model() -> None:
    """
    Using the wind tunnel data supplied, construct the following surrogate model:
    𝑐total =𝑓(𝐻1,𝐻2)
    evaluate 𝑐total at the H1, H2 values supplied in the Q1 tab of “ML_test.xlsx”.
    """
    inference_data_source = "Q1"

    data = load_ml_test_data(xlsx_data_filename=XLSX_DATA_FILENAME, directory=RAW_DATA_DIR)
    predictions = features.modeling(
        data=data, input_data_cols=["H1", "H2"], inference_data_source=inference_data_source
    )

    # report
    process_workbook(
        sheet_name=inference_data_source,
        data_to_write=[[value] for value in predictions],
        column_map=["C"],
    )


def H1_H2_for_lowest_ctotal() -> None:
    """
    identify the values of H1 and H2 that result in the lowest ctotal using the surrogate model
    """

    data = load_ml_test_data(xlsx_data_filename=XLSX_DATA_FILENAME, directory=RAW_DATA_DIR)
    optimal_H = features.get_optimal_H(data=data, input_data_cols=["H1", "H2"])

    # report
    process_workbook(
        sheet_name="Q2",
        data_to_write=[optimal_H.values()],
        column_map=["A", "B", "C"],
        start_row=2,
    )


def surrogate_model_from_discrete_points() -> None:
    """
    Using the wind tunnel data supplied, construct the following surrogate model:
    𝑐total = g(p0, ..., p111)
    evaluate 𝑐total at the p0, ..., p111 values supplied in the Q3 tab of “ML_test.xlsx”.
    """
    inference_data_source = "Q3"
    recorded_points = [f"p{i}" for i in range(0, 112)]

    data = load_ml_test_data(xlsx_data_filename=XLSX_DATA_FILENAME, directory=RAW_DATA_DIR)
    predictions = features.modeling(
        data=data, input_data_cols=recorded_points, inference_data_source=inference_data_source
    )

    # report
    process_workbook(
        sheet_name=inference_data_source,
        data_to_write=[[value] for value in predictions],
        column_map=["DI"],
    )


def surrogate_model_optimal_pressure_sensors() -> None:
    """
    picking subset of 20 pressure sensors for the wind tunnel model training with optimal accuracy
    """

    recorded_points = [f"p{i}" for i in range(0, 112)]
    inference_data_source = "Q4"

    data = load_ml_test_data(xlsx_data_filename=XLSX_DATA_FILENAME, directory=RAW_DATA_DIR)
    predictions = features.model_optimization(
        data=data, input_data_cols=recorded_points, inference_data_source=inference_data_source
    )

    # report
    process_workbook(
        sheet_name=inference_data_source,
        data_to_write=[[value] for value in predictions],
        column_map=["DI"],
    )


def minimize_wind_tunnel_time() -> None:
    """
    minimize the wind tunnel time of a given run while still attempting to
    cover the (H1, H2) space as well as possible
    """

    n_samples = 20
    data = load_ml_test_data(xlsx_data_filename=XLSX_DATA_FILENAME, directory=RAW_DATA_DIR)
    bounds = [(0.01, 0.2), (0.01, 0.344)]

    optimal_n_samples = features.find_n_samples_iteratively(
        data=data, input_data_cols=["H1", "H2"], bounds=bounds, n_samples=n_samples
    )

    optimal_experiment = features.wind_tunnel_time_optimization(
        n_samples=optimal_n_samples, bounds=bounds
    )

    # report
    process_workbook(
        sheet_name="Q5",
        data_to_write=[i.values() for i in optimal_experiment],
        column_map=["A", "B", "C"],
        start_row=2,
    )
